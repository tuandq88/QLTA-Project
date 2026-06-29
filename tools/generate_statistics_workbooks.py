"""Sinh metadata giao diện từ các workbook gốc trong bieu_mau/.

Script chỉ đọc workbook nguồn và tạo JSON cho frontend. Các số liệu mẫu nằm dưới
dòng đánh số cột được để trống để giao diện chỉ nhận số liệu đã tính từ API.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Color
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "bieu_mau"
OUTPUT = ROOT / "frontend" / "src" / "data" / "statistics-workbooks.json"

FORM_CODES = {
    "Hình sự sơ thẩm.xlsx": "HS_ST_1A",
    "Hình sự phúc thẩm.xlsx": "HS_PT_1B",
    "Dân sự sơ thẩm.xlsx": "DS_ST_2A",
    "Dân sự Phúc thẩm.xlsx": "DS_PT_2B",
    "Hôn nhân gia đình sơ thẩm.xlsx": "HNGD_ST_3A",
    "Hôn nhân gia đình phúc thẩm.xlsx": "HNGD_PT_3B",
    "Kinh doanh thương mại sơ thẩm.xlsx": "KDTM_ST_4A",
    "Kinh doanh thương mại phúc thẩm.xlsx": "KDTM_PT_4B",
    "Lao động sơ thẩm.xlsx": "LD_ST_5A",
    "Lao động phúc thẩm.xlsx": "LD_PT_5B",
    "Hành chính sơ thẩm.xlsx": "HC_ST_6A",
    "Hành chính phúc thẩm.xlsx": "HC_PT_6B",
}

CASE_TYPES = [
    ("Hình sự", "hinh-su"),
    ("Dân sự", "dan-su"),
    ("Hôn nhân gia đình", "hon-nhan-gia-dinh"),
    ("Kinh doanh thương mại", "kinh-doanh-thuong-mai"),
    ("Lao động", "lao-dong"),
    ("Hành chính", "hanh-chinh"),
]

STAGES = [
    ("sơ thẩm", "so-tham", "Sơ thẩm"),
    ("phúc thẩm", "phuc-tham", "Phúc thẩm"),
    ("giám đốc thẩm", "giam-doc-tham", "Giám đốc thẩm"),
    ("tái thẩm", "tai-tham", "Tái thẩm"),
]


def rgb(color: Color | None) -> str | None:
    if not color or color.type != "rgb" or not color.rgb:
        return None
    value = str(color.rgb)
    return f"#{value[-6:]}" if value[-6:] != "000000" or value in {"FF000000", "00000000"} else "#000000"


def border_side(side: Any) -> dict[str, str] | None:
    if not side or not side.style:
        return None
    widths = {"hair": "1px", "thin": "1px", "medium": "2px", "thick": "3px", "double": "3px"}
    styles = {"dashed": "dashed", "dotted": "dotted", "double": "double"}
    return {
        "width": widths.get(side.style, "1px"),
        "style": styles.get(side.style, "solid"),
        "color": rgb(side.color) or "#000000",
    }


def plain_value(value: Any) -> str | int | float | None:
    if value is None:
        return None
    if isinstance(value, (str, int, float)):
        return value
    return str(value)


def workbook_identity(filename: str) -> tuple[str, str, str, str]:
    lower = filename.casefold()
    case_label = case_id = ""
    for label, identifier in CASE_TYPES:
        if lower.startswith(label.casefold()):
            case_label, case_id = label, identifier
            break
    stage_label = stage_id = ""
    for needle, identifier, label in STAGES:
        if needle in lower:
            stage_label, stage_id = label, identifier
            break
    if not case_id or not stage_id:
        raise ValueError(f"Không phân loại được workbook: {filename}")
    return f"{case_id}-{stage_id}", case_label, stage_id, stage_label


def find_number_row(ws: Any) -> int:
    best_row, best_run = 1, 0
    for row in range(1, ws.max_row + 1):
        run = 0
        for column in range(1, ws.max_column + 1):
            value = ws.cell(row, column).value
            if isinstance(value, int) and value == column:
                run += 1
            elif column > 2:
                break
        if run > best_run:
            best_row, best_run = row, run
    return best_row


def find_total_row(ws: Any, number_row: int) -> int:
    total = ws.max_row
    for row in range(number_row + 1, ws.max_row + 1):
        value = ws.cell(row, 1).value
        if isinstance(value, str) and value.strip().casefold() == "tổng cộng":
            total = row
    return total


def convert(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False, read_only=False)
    ws = workbook.active
    identifier, case_label, stage_id, stage_label = workbook_identity(path.name)
    number_row = find_number_row(ws)
    total_row = find_total_row(ws, number_row)
    merged_children: set[tuple[int, int]] = set()
    merge_anchors: dict[tuple[int, int], tuple[int, int]] = {}
    for merged in ws.merged_cells.ranges:
        merge_anchors[(merged.min_row, merged.min_col)] = (
            merged.max_row - merged.min_row + 1,
            merged.max_col - merged.min_col + 1,
        )
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                if (row, column) != (merged.min_row, merged.min_col):
                    merged_children.add((row, column))

    cells: dict[str, Any] = {}
    styles: list[dict[str, Any]] = []
    style_indexes: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            if (row, column) in merged_children:
                continue
            cell = ws.cell(row, column)
            if isinstance(cell, MergedCell):
                continue
            value = plain_value(cell.value)
            if row > number_row and isinstance(value, (int, float)):
                value = None
            fill = rgb(cell.fill.fgColor) if cell.fill and cell.fill.fill_type else None
            font_color = rgb(cell.font.color)
            rowspan, colspan = merge_anchors.get((row, column), (1, 1))
            style = {
                "background": fill,
                "color": font_color,
                "fontFamily": cell.font.name,
                "fontSize": cell.font.sz,
                "fontWeight": 700 if cell.font.bold else 400,
                "fontStyle": "italic" if cell.font.italic else None,
                "textDecoration": "underline" if cell.font.underline else None,
                "horizontal": cell.alignment.horizontal,
                "vertical": cell.alignment.vertical,
                "wrapText": bool(cell.alignment.wrap_text),
                "rotation": cell.alignment.text_rotation or 0,
                "borderTop": border_side(cell.border.top),
                "borderRight": border_side(cell.border.right),
                "borderBottom": border_side(cell.border.bottom),
                "borderLeft": border_side(cell.border.left),
            }
            compact_style = {key: item for key, item in style.items() if item is not None}
            style_key = json.dumps(compact_style, ensure_ascii=False, sort_keys=True)
            if style_key not in style_indexes:
                style_indexes[style_key] = len(styles)
                styles.append(compact_style)
            compact_cell: dict[str, Any] = {"s": style_indexes[style_key]}
            if value is not None:
                compact_cell["v"] = value
            if rowspan != 1:
                compact_cell["rs"] = rowspan
            if colspan != 1:
                compact_cell["cs"] = colspan
            cells[f"{row}:{column}"] = compact_cell

    columns = []
    for column in range(1, ws.max_column + 1):
        letter = get_column_letter(column)
        dimension = ws.column_dimensions[letter]
        width = dimension.width if dimension.width is not None else 13
        columns.append([round(max(18, width * 7 + 5), 2), bool(dimension.hidden)])

    rows = []
    for row in range(1, ws.max_row + 1):
        dimension = ws.row_dimensions[row]
        height = dimension.height if dimension.height is not None else 15
        rows.append([round(max(10, height * 96 / 72), 2), bool(dimension.hidden)])

    return {
        "id": identifier,
        "title": f"{case_label} {stage_label.lower()}",
        "caseType": case_label,
        "stage": stage_id,
        "stageLabel": stage_label,
        "fileName": path.name,
        "sheetName": ws.title,
        "formCode": FORM_CODES.get(path.name),
        "numberRow": number_row,
        "totalRow": total_row,
        "maxRow": ws.max_row,
        "maxColumn": ws.max_column,
        "columns": columns,
        "rows": rows,
        "styles": styles,
        "cells": cells,
    }


def main() -> None:
    workbooks = [convert(path) for path in sorted(SOURCE_DIR.glob("*.xlsx")) if not path.name.startswith("~$")]
    workbooks.sort(key=lambda item: (
        [label for label, _ in CASE_TYPES].index(item["caseType"]),
        [identifier for _, identifier, _ in STAGES].index(item["stage"]),
    ))
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(json.dumps(workbooks, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Generated {len(workbooks)} workbooks at {OUTPUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
