#!/usr/bin/env python
"""Generate an Excel report for criminal first-instance cases in a period.

The script reads PostgreSQL through psql so it does not require a Python
database driver. It writes xlsx output with openpyxl.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import shutil
import subprocess
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = REPO_ROOT / "tests" / "criminal_first_instance_cases_2026_06_01_to_2026_06_10.xlsx"
SCRIPT_RELATIVE_PATH = "tests/generate_criminal_first_instance_cases_report_2026_06_01_to_2026_06_10.py"
DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


DETAIL_COLUMNS = [
    "case_id",
    "case_code",
    "case_number",
    "case_type",
    "case_type_name",
    "case_group",
    "case_group_name",
    "procedure_law",
    "acceptance_date",
    "closed_date",
    "current_stage",
    "case_status",
    "resolution_status",
    "summary",
    "court_id",
    "court_code",
    "court_name",
    "procuracy_name",
    "indictment_number",
    "indictment_date",
    "investigation_agency",
    "dossier_received_date",
    "has_civil_claim",
    "has_minor_defendant",
    "has_legal_aid",
    "trial_panel_type",
    "defendant_count",
    "minor_defendant_count",
    "defendant_names",
    "crime_names",
    "penal_code_articles",
    "clause_points",
    "crime_severities",
    "sentence_types",
    "fine_amount_total",
    "imprisonment_years_total",
    "suspended_years_total",
    "probation_years_total",
    "additional_penalties",
    "civil_liabilities",
    "decision_numbers",
    "decision_dates",
    "decision_result_codes",
    "decision_result_summaries",
    "presiding_judges",
    "panel_judges",
    "panel_judge_count",
    "hearing_clerks",
    "hearing_clerk_count",
    "in_accepted_period",
    "in_resolved_period",
    "unresolved_as_of_date",
    "matched_groups",
]


def load_dotenv(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8-sig").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        name, value = line.split("=", 1)
        name = name.strip()
        value = value.strip().strip('"').strip("'")
        os.environ.setdefault(name, value)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate criminal first-instance case Excel report.")
    parser.add_argument("--database", default=os.environ.get("PGDATABASE", "qlta_schema_merge_test"))
    parser.add_argument("--from-date", default="2026-06-01")
    parser.add_argument("--to-date", default="2026-06-10")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    return parser.parse_args()


def validate_date(value: str, name: str) -> str:
    if not DATE_RE.match(value):
        raise SystemExit(f"{name} must use YYYY-MM-DD format, got: {value}")
    return value


def psql_base_args(database: str) -> list[str]:
    psql = shutil.which("psql")
    if not psql:
        raise SystemExit("Cannot find psql in PATH.")
    return [
        psql,
        "-h",
        os.environ.get("PGHOST", "localhost"),
        "-p",
        os.environ.get("PGPORT", "5432"),
        "-U",
        os.environ.get("PGUSER", "postgres"),
        "-d",
        database,
        "-v",
        "ON_ERROR_STOP=1",
        "-q",
    ]


def psql_env() -> dict[str, str]:
    env = os.environ.copy()
    env.setdefault("PGCLIENTENCODING", "UTF8")
    return env


def run_psql_scalar(database: str, sql: str) -> str:
    result = subprocess.run(
        psql_base_args(database) + ["-t", "-A", "-c", sql],
        cwd=REPO_ROOT,
        text=True,
        encoding="utf-8",
        errors="replace",
        capture_output=True,
        env=psql_env(),
        check=False,
    )
    if result.returncode != 0:
        raise SystemExit(result.stderr.strip() or result.stdout.strip())
    return result.stdout.strip()


def run_copy_csv(database: str, query: str) -> list[dict[str, str]]:
    sql = f"COPY ({query}) TO STDOUT WITH CSV HEADER"
    with tempfile.NamedTemporaryFile("w", suffix=".sql", encoding="utf-8", delete=False) as handle:
        handle.write(sql)
        handle.write(";\n")
        sql_path = Path(handle.name)
    try:
        result = subprocess.run(
            psql_base_args(database) + ["-f", str(sql_path)],
            cwd=REPO_ROOT,
            text=True,
            encoding="utf-8",
            errors="replace",
            capture_output=True,
            env=psql_env(),
            check=False,
        )
    finally:
        sql_path.unlink(missing_ok=True)
    if result.returncode != 0:
        raise SystemExit(result.stderr.strip() or result.stdout.strip())
    return list(csv.DictReader(result.stdout.splitlines()))


def check_database_contract(database: str) -> None:
    case_files_exists = run_psql_scalar(database, "SELECT to_regclass('public.case_files') IS NOT NULL;")
    if case_files_exists.lower() != "t":
        raise SystemExit("Required table public.case_files was not found.")

    criminal_exists = run_psql_scalar(
        database,
        """
        SELECT EXISTS (
            SELECT 1
            FROM dm_categories c
            JOIN dm_category_items i ON i.category_id = c.category_id
            WHERE c.category_code = 'case_type'
              AND i.item_code = 'criminal'
        ) OR EXISTS (
            SELECT 1 FROM case_files WHERE lower(case_type::text) = 'criminal'
        );
        """,
    )
    if criminal_exists.lower() != "t":
        raise SystemExit("Cannot identify criminal case type from category case_type or case_files.case_type.")

    first_instance_exists = run_psql_scalar(
        database,
        """
        SELECT EXISTS (
            SELECT 1
            FROM dm_categories c
            JOIN dm_category_items i ON i.category_id = c.category_id
            WHERE c.category_code = 'case_group'
              AND i.item_code = 'SO_THAM'
        ) OR EXISTS (
            SELECT 1 FROM case_files WHERE case_group = 'SO_THAM'
        );
        """,
    )
    if first_instance_exists.lower() != "t":
        raise SystemExit("Cannot identify first-instance case group SO_THAM.")


def base_ctes(from_date: str, to_date: str) -> str:
    return f"""
WITH params AS (
    SELECT DATE '{from_date}' AS from_date, DATE '{to_date}' AS to_date
),
case_type_catalog AS (
    SELECT i.item_id, i.item_code, i.item_name
    FROM dm_categories c
    JOIN dm_category_items i ON i.category_id = c.category_id
    WHERE c.category_code = 'case_type'
),
case_group_catalog AS (
    SELECT i.item_id, i.item_code, i.item_name
    FROM dm_categories c
    JOIN dm_category_items i ON i.category_id = c.category_id
    WHERE c.category_code = 'case_group'
),
defendant_stats AS (
    SELECT
        ccd.case_id,
        count(d.defendant_id)::integer AS defendant_count,
        count(d.defendant_id) FILTER (WHERE d.is_minor IS TRUE)::integer AS minor_defendant_count,
        string_agg(DISTINCT d.full_name, '; ' ORDER BY d.full_name) AS defendant_names
    FROM criminal_case_details ccd
    LEFT JOIN defendants d ON d.criminal_detail_id = ccd.criminal_detail_id
    GROUP BY ccd.case_id
),
charge_stats AS (
    SELECT
        ccd.case_id,
        string_agg(DISTINCT ch.crime_name, '; ' ORDER BY ch.crime_name) FILTER (WHERE ch.crime_name IS NOT NULL) AS crime_names,
        string_agg(DISTINCT ch.penal_code_article, '; ' ORDER BY ch.penal_code_article) FILTER (WHERE ch.penal_code_article IS NOT NULL) AS penal_code_articles,
        string_agg(DISTINCT ch.clause_point, '; ' ORDER BY ch.clause_point) FILTER (WHERE ch.clause_point IS NOT NULL) AS clause_points,
        string_agg(DISTINCT ch.crime_severity, '; ' ORDER BY ch.crime_severity) FILTER (WHERE ch.crime_severity IS NOT NULL) AS crime_severities
    FROM criminal_case_details ccd
    JOIN defendants d ON d.criminal_detail_id = ccd.criminal_detail_id
    LEFT JOIN charges ch ON ch.defendant_id = d.defendant_id
    GROUP BY ccd.case_id
),
sentence_stats AS (
    SELECT
        ccd.case_id,
        string_agg(DISTINCT s.sentence_type, '; ' ORDER BY s.sentence_type) FILTER (WHERE s.sentence_type IS NOT NULL) AS sentence_types,
        sum(s.fine_amount) AS fine_amount_total,
        sum(s.imprisonment_years) AS imprisonment_years_total,
        sum(s.suspended_years) AS suspended_years_total,
        sum(s.probation_years) AS probation_years_total,
        string_agg(DISTINCT s.additional_penalty, '; ' ORDER BY s.additional_penalty) FILTER (WHERE s.additional_penalty IS NOT NULL) AS additional_penalties,
        string_agg(DISTINCT s.civil_liability, '; ' ORDER BY s.civil_liability) FILTER (WHERE s.civil_liability IS NOT NULL) AS civil_liabilities
    FROM criminal_case_details ccd
    JOIN defendants d ON d.criminal_detail_id = ccd.criminal_detail_id
    LEFT JOIN sentences s ON s.defendant_id = d.defendant_id
    GROUP BY ccd.case_id
),
decision_stats AS (
    SELECT
        case_id,
        string_agg(DISTINCT decision_number, '; ' ORDER BY decision_number) FILTER (WHERE decision_number IS NOT NULL) AS decision_numbers,
        string_agg(DISTINCT decision_date::text, '; ' ORDER BY decision_date::text) FILTER (WHERE decision_date IS NOT NULL) AS decision_dates,
        string_agg(DISTINCT result_code, '; ' ORDER BY result_code) FILTER (WHERE result_code IS NOT NULL) AS decision_result_codes,
        string_agg(DISTINCT result_summary, '; ' ORDER BY result_summary) FILTER (WHERE result_summary IS NOT NULL) AS decision_result_summaries
    FROM decisions
    GROUP BY case_id
),
member_stats AS (
    SELECT
        chm.case_id,
        string_agg(cs.full_name, '; ' ORDER BY chm.member_order, cs.full_name) FILTER (WHERE chm.role_code = 'PRESIDING_JUDGE') AS presiding_judges,
        string_agg(cs.full_name, '; ' ORDER BY chm.member_order, cs.full_name) FILTER (WHERE chm.role_code = 'PANEL_JUDGE') AS panel_judges,
        count(*) FILTER (WHERE chm.role_code = 'PANEL_JUDGE')::integer AS panel_judge_count,
        string_agg(cs.full_name, '; ' ORDER BY chm.member_order, cs.full_name) FILTER (WHERE chm.role_code = 'HEARING_CLERK') AS hearing_clerks,
        count(*) FILTER (WHERE chm.role_code = 'HEARING_CLERK')::integer AS hearing_clerk_count
    FROM case_hearing_members chm
    JOIN court_staff cs ON cs.staff_id = chm.staff_id
    GROUP BY chm.case_id
),
base_cases AS (
    SELECT
        cf.case_id,
        cf.case_code,
        cf.case_number,
        cf.case_type::text AS case_type,
        coalesce(ct.item_name, ct_text.item_name) AS case_type_name,
        cf.case_group,
        coalesce(cg.item_name, cg_text.item_name) AS case_group_name,
        cf.procedure_law,
        cf.acceptance_date,
        cf.closed_date,
        cf.current_stage,
        cf.case_status::text AS case_status,
        cf.resolution_status,
        cf.summary,
        cf.court_id,
        courts.court_code,
        courts.court_name,
        ccd.procuracy_name,
        ccd.indictment_number,
        ccd.indictment_date,
        ccd.investigation_agency,
        ccd.dossier_received_date,
        ccd.has_civil_claim,
        ccd.has_minor_defendant,
        ccd.has_legal_aid,
        ccd.trial_panel_type,
        coalesce(ds.defendant_count, 0) AS defendant_count,
        coalesce(ds.minor_defendant_count, 0) AS minor_defendant_count,
        ds.defendant_names,
        charge_stats.crime_names,
        charge_stats.penal_code_articles,
        charge_stats.clause_points,
        charge_stats.crime_severities,
        sentence_stats.sentence_types,
        sentence_stats.fine_amount_total,
        sentence_stats.imprisonment_years_total,
        sentence_stats.suspended_years_total,
        sentence_stats.probation_years_total,
        sentence_stats.additional_penalties,
        sentence_stats.civil_liabilities,
        decision_stats.decision_numbers,
        decision_stats.decision_dates,
        decision_stats.decision_result_codes,
        decision_stats.decision_result_summaries,
        member_stats.presiding_judges,
        member_stats.panel_judges,
        coalesce(member_stats.panel_judge_count, 0) AS panel_judge_count,
        member_stats.hearing_clerks,
        coalesce(member_stats.hearing_clerk_count, 0) AS hearing_clerk_count,
        (cf.acceptance_date BETWEEN p.from_date AND p.to_date) AS in_accepted_period,
        (cf.closed_date BETWEEN p.from_date AND p.to_date) AS in_resolved_period,
        (cf.acceptance_date <= p.to_date AND (cf.closed_date IS NULL OR cf.closed_date > p.to_date)) AS unresolved_as_of_date
    FROM case_files cf
    CROSS JOIN params p
    LEFT JOIN courts ON courts.court_id = cf.court_id
    LEFT JOIN case_type_catalog ct ON ct.item_id = cf.case_type_id
    LEFT JOIN case_type_catalog ct_text ON ct_text.item_code = cf.case_type::text
    LEFT JOIN case_group_catalog cg ON cg.item_id = cf.case_group_id
    LEFT JOIN case_group_catalog cg_text ON cg_text.item_code = cf.case_group
    LEFT JOIN criminal_case_details ccd ON ccd.case_id = cf.case_id
    LEFT JOIN defendant_stats ds ON ds.case_id = cf.case_id
    LEFT JOIN charge_stats ON charge_stats.case_id = cf.case_id
    LEFT JOIN sentence_stats ON sentence_stats.case_id = cf.case_id
    LEFT JOIN decision_stats ON decision_stats.case_id = cf.case_id
    LEFT JOIN member_stats ON member_stats.case_id = cf.case_id
    WHERE cf.acceptance_date <= p.to_date
      AND (
            ct.item_code = 'criminal'
         OR ct_text.item_code = 'criminal'
         OR cf.case_type::text = 'criminal'
      )
      AND (
            cg.item_code = 'SO_THAM'
         OR cg_text.item_code = 'SO_THAM'
         OR cf.case_group = 'SO_THAM'
      )
),
enriched_cases AS (
    SELECT
        base_cases.*,
        array_to_string(
            ARRAY_REMOVE(ARRAY[
                CASE WHEN in_accepted_period THEN 'THU_LY_TRONG_KY' END,
                CASE WHEN in_resolved_period THEN 'GIAI_QUYET_TRONG_KY' END,
                CASE WHEN unresolved_as_of_date THEN 'CON_LAI_DEN_{to_date.replace('-', '_')}' END
            ], NULL),
            ', '
        ) AS matched_groups
    FROM base_cases
)
"""


def detail_query(from_date: str, to_date: str, where_clause: str) -> str:
    columns = ",\n        ".join(DETAIL_COLUMNS)
    return f"""
{base_ctes(from_date, to_date)}
SELECT
        {columns}
FROM enriched_cases
WHERE {where_clause}
ORDER BY acceptance_date NULLS LAST, case_number NULLS LAST, case_code NULLS LAST, case_id
"""


def validation_query(from_date: str, to_date: str) -> str:
    return f"""
WITH params AS (
    SELECT DATE '{from_date}' AS from_date, DATE '{to_date}' AS to_date
),
case_type_catalog AS (
    SELECT i.item_id, i.item_code
    FROM dm_categories c
    JOIN dm_category_items i ON i.category_id = c.category_id
    WHERE c.category_code = 'case_type'
),
case_group_catalog AS (
    SELECT i.item_id, i.item_code
    FROM dm_categories c
    JOIN dm_category_items i ON i.category_id = c.category_id
    WHERE c.category_code = 'case_group'
),
target_cases AS (
    SELECT cf.*
    FROM case_files cf
    CROSS JOIN params p
    LEFT JOIN case_type_catalog ct ON ct.item_id = cf.case_type_id
    LEFT JOIN case_type_catalog ct_text ON ct_text.item_code = cf.case_type::text
    LEFT JOIN case_group_catalog cg ON cg.item_id = cf.case_group_id
    LEFT JOIN case_group_catalog cg_text ON cg_text.item_code = cf.case_group
    WHERE cf.acceptance_date <= p.to_date
      AND (ct.item_code = 'criminal' OR ct_text.item_code = 'criminal' OR cf.case_type::text = 'criminal')
      AND (cg.item_code = 'SO_THAM' OR cg_text.item_code = 'SO_THAM' OR cf.case_group = 'SO_THAM')
),
member_counts AS (
    SELECT
        chm.case_id,
        count(*) FILTER (WHERE chm.role_code = 'PRESIDING_JUDGE') AS presiding_count,
        count(*) FILTER (WHERE chm.role_code = 'HEARING_CLERK') AS clerk_count,
        count(*) FILTER (WHERE chm.role_code = 'PANEL_JUDGE') AS panel_count
    FROM case_hearing_members chm
    GROUP BY chm.case_id
)
SELECT 'MISSING_CASE_GROUP' AS validation_code,
       'Số vụ án hình sự thiếu case_group/cấp xét xử' AS validation_name,
       count(*)::text AS value,
       'warning' AS severity,
       'Lọc trong case_files có case_type=criminal, acceptance_date<=to_date.' AS note
FROM case_files cf CROSS JOIN params p
WHERE cf.acceptance_date <= p.to_date
  AND cf.case_type::text = 'criminal'
  AND (cf.case_group IS NULL OR cf.case_group_id IS NULL)
UNION ALL
SELECT 'MISSING_CASE_TYPE', 'Số vụ án thiếu case_type', count(*)::text, 'warning',
       'Đếm trên case_files có acceptance_date<=to_date.'
FROM case_files cf CROSS JOIN params p
WHERE cf.acceptance_date <= p.to_date
  AND cf.case_type IS NULL
UNION ALL
SELECT 'CLOSED_BEFORE_ACCEPTANCE', 'Số vụ án có closed_date < acceptance_date', count(*)::text, 'error',
       'Đếm trên toàn bộ case_files.'
FROM case_files
WHERE closed_date IS NOT NULL AND acceptance_date IS NOT NULL AND closed_date < acceptance_date
UNION ALL
SELECT 'CASE_TYPE_CATEGORY_NOT_JOINED', 'Số vụ án không join được danh mục case_type', count(*)::text, 'warning',
       'Đếm trên case_files có acceptance_date<=to_date.'
FROM case_files cf
CROSS JOIN params p
LEFT JOIN case_type_catalog ct ON ct.item_id = cf.case_type_id
LEFT JOIN case_type_catalog ct_text ON ct_text.item_code = cf.case_type::text
WHERE cf.acceptance_date <= p.to_date
  AND ct.item_id IS NULL
  AND ct_text.item_id IS NULL
UNION ALL
SELECT 'CASE_GROUP_CATEGORY_NOT_JOINED', 'Số vụ án không join được danh mục case_group', count(*)::text, 'warning',
       'Đếm trên case_files có acceptance_date<=to_date.'
FROM case_files cf
CROSS JOIN params p
LEFT JOIN case_group_catalog cg ON cg.item_id = cf.case_group_id
LEFT JOIN case_group_catalog cg_text ON cg_text.item_code = cf.case_group
WHERE cf.acceptance_date <= p.to_date
  AND cg.item_id IS NULL
  AND cg_text.item_id IS NULL
UNION ALL
SELECT 'NO_PRESIDING_JUDGE', 'Số vụ án không có thẩm phán chủ tọa', count(*)::text, 'warning',
       'Áp dụng cho hình sự sơ thẩm đến to_date nếu có case_hearing_members.'
FROM target_cases tc
LEFT JOIN member_counts mc ON mc.case_id = tc.case_id
WHERE coalesce(mc.presiding_count, 0) = 0
UNION ALL
SELECT 'NO_HEARING_CLERK', 'Số vụ án không có thư ký phiên tòa', count(*)::text, 'warning',
       'Áp dụng cho hình sự sơ thẩm đến to_date nếu có case_hearing_members.'
FROM target_cases tc
LEFT JOIN member_counts mc ON mc.case_id = tc.case_id
WHERE coalesce(mc.clerk_count, 0) = 0
UNION ALL
SELECT 'FIRST_INSTANCE_PANEL_GT_1', 'Số vụ án sơ thẩm có hơn 1 thẩm phán thành viên hội đồng', count(*)::text, 'warning',
       'Theo skill_hearing_members, sơ thẩm cho phép 0 hoặc 1 PANEL_JUDGE.'
FROM target_cases tc
LEFT JOIN member_counts mc ON mc.case_id = tc.case_id
WHERE coalesce(mc.panel_count, 0) > 1
"""


def write_rows(ws, rows: list[dict[str, str]], headers: list[str], header_fill) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for value in (row.get(header, "") for row in rows[:200]):
            max_len = max(max_len, len(str(value or "")))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 42)


def write_key_values(ws, rows: Iterable[tuple[str, str]], header_fill) -> None:
    ws.append(["field", "value"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for key, value in rows:
        ws.append([key, value])
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 110
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_workbook(
    database: str,
    from_date: str,
    to_date: str,
    output_path: Path,
    sheets: dict[str, list[dict[str, str]]],
    validation_rows: list[dict[str, str]],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")

    readme = wb.create_sheet("README")
    write_key_values(
        readme,
        [
            ("Mục đích", "Danh sách chi tiết các vụ án hình sự sơ thẩm theo 3 nhóm: thụ lý trong kỳ, giải quyết trong kỳ, còn lại đến ngày báo cáo."),
            ("Database", database),
            ("Kỳ thống kê", f"{from_date} đến {to_date}"),
            ("as_of_date", to_date),
            ("Loại án", "Hình sự - case_type/category code criminal"),
            ("Cấp xét xử", "Sơ thẩm - case_group/category code SO_THAM"),
            ("Quy tắc án thụ lý", f"acceptance_date BETWEEN DATE '{from_date}' AND DATE '{to_date}'"),
            ("Quy tắc án giải quyết", f"closed_date BETWEEN DATE '{from_date}' AND DATE '{to_date}'"),
            ("Quy tắc án còn lại", f"acceptance_date <= DATE '{to_date}' AND (closed_date IS NULL OR closed_date > DATE '{to_date}')"),
            ("Không đếm trùng", "SUMMARY có TONG_HOP_KHONG_TRUNG từ ALL_MATCHED_CASES; các sheet chi tiết theo nhóm có thể lặp cùng case_id nếu vụ án thuộc nhiều nhóm."),
            ("Ngày tạo file", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Script tạo file", SCRIPT_RELATIVE_PATH),
            ("Skill đã áp dụng", "skill_statistics_mapping, skill_thong_ke_hinh_su, skill_trial_level_classification, skill_hearing_members"),
        ],
        header_fill,
    )

    summary_rows = [
        {
            "metric_code": "THU_LY_TRONG_KY",
            "metric_name": "Án thụ lý trong kỳ",
            "total_cases": str(len(sheets["ACCEPTED_IN_PERIOD"])),
            "note": "Không cộng trùng với các nhóm khác.",
        },
        {
            "metric_code": "GIAI_QUYET_TRONG_KY",
            "metric_name": "Án giải quyết trong kỳ",
            "total_cases": str(len(sheets["RESOLVED_IN_PERIOD"])),
            "note": "Dựa trên closed_date trong kỳ.",
        },
        {
            "metric_code": f"CON_LAI_DEN_{to_date.replace('-', '_')}",
            "metric_name": f"Án còn lại/chưa giải quyết đến {to_date}",
            "total_cases": str(len(sheets["UNRESOLVED_AS_OF_DATE"])),
            "note": "Bao gồm vụ án có closed_date sau as_of_date.",
        },
        {
            "metric_code": "TONG_HOP_KHONG_TRUNG",
            "metric_name": "Tổng hợp unique toàn bộ vụ án thuộc ít nhất một nhóm",
            "total_cases": str(len(sheets["ALL_MATCHED_CASES"])),
            "note": "Một case_id chỉ tính một lần.",
        },
    ]
    ws = wb.create_sheet("SUMMARY")
    write_rows(ws, summary_rows, ["metric_code", "metric_name", "total_cases", "note"], header_fill)

    for sheet_name in [
        "ACCEPTED_IN_PERIOD",
        "RESOLVED_IN_PERIOD",
        "UNRESOLVED_AS_OF_DATE",
        "ALL_MATCHED_CASES",
    ]:
        ws = wb.create_sheet(sheet_name)
        write_rows(ws, sheets[sheet_name], DETAIL_COLUMNS, header_fill)

    ws = wb.create_sheet("VALIDATION")
    write_rows(ws, validation_rows, ["validation_code", "validation_name", "value", "severity", "note"], header_fill)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    load_dotenv(REPO_ROOT / ".env.local")
    args = parse_args()
    database = args.database
    from_date = validate_date(args.from_date, "--from-date")
    to_date = validate_date(args.to_date, "--to-date")
    output_path = Path(args.output)
    if not output_path.is_absolute():
        output_path = REPO_ROOT / output_path

    check_database_contract(database)

    sheets = {
        "ACCEPTED_IN_PERIOD": run_copy_csv(database, detail_query(from_date, to_date, "in_accepted_period IS TRUE")),
        "RESOLVED_IN_PERIOD": run_copy_csv(database, detail_query(from_date, to_date, "in_resolved_period IS TRUE")),
        "UNRESOLVED_AS_OF_DATE": run_copy_csv(database, detail_query(from_date, to_date, "unresolved_as_of_date IS TRUE")),
        "ALL_MATCHED_CASES": run_copy_csv(
            database,
            detail_query(
                from_date,
                to_date,
                "in_accepted_period IS TRUE OR in_resolved_period IS TRUE OR unresolved_as_of_date IS TRUE",
            ),
        ),
    }
    validation_rows = run_copy_csv(database, validation_query(from_date, to_date))

    build_workbook(database, from_date, to_date, output_path, sheets, validation_rows)

    print(f"Created: {output_path}")
    for sheet_name, rows in sheets.items():
        print(f"{sheet_name}: {len(rows)} rows")
    warning_count = sum(1 for row in validation_rows if row.get("severity") in {"warning", "error"} and row.get("value") not in {"0", "", None})
    print(f"VALIDATION: {len(validation_rows)} rows, {warning_count} warning/error metrics with value > 0")
    return 0


if __name__ == "__main__":
    sys.exit(main())
